import streamlit as st
import pandas as pd
import pulp
import io
import hashlib
import time
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Seitenkonfiguration
st.set_page_config(page_title="4D Zielwert-Optimierer", page_icon="📊", layout="wide")

st.title("🧮 Fast 4D Zielwert-Optimierer")
st.write("Optimierte Version für größere Datenmengen (z.B. 60+ Zeilen) mit automatischer Zeitbegrenzung.")

# ==========================================
# 1. TEMPLATE-DOWNLOAD FÜR DIE STRUKTUR (FIXED)
# ==========================================
st.sidebar.header("1. Struktur-Vorlage")

# Hier wurde die Liste für die Vorlage korrekt erstellt (60 Zeilen mit 0 als Platzhalter)
template_df = pd.DataFrame({
    "Objektname": [f"Objekt_{i}" for i in range(1, 61)],
    "Jahr_1": [0] * 60,
    "Jahr_2": [0] * 60,
    "Jahr_3": [0] * 60,
    "Jahr_4": [0] * 60
})

zielwerte_template_df = pd.DataFrame({
    "Jahr": ["Jahr 1", "Jahr 2", "Jahr 3", "Jahr 4"],
    "Zielwert": [None, None, None, None]
})

buffer_template = io.BytesIO()
with pd.ExcelWriter(buffer_template, engine='openpyxl') as writer:
    template_df.to_excel(writer, index=False, sheet_name="Daten")
    zielwerte_template_df.to_excel(writer, index=False, sheet_name="Zielwerte")

st.sidebar.download_button(
    label="📄 Excel-Vorlage herunterladen",
    data=buffer_template.getvalue(),
    file_name="struktur_vorlage_4d.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ==========================================
# 2. DATEI-UPLOAD & ZIELWERTE
# ==========================================
uploaded_file = st.file_uploader("Wählen Sie Ihre ausgefüllte Excel-Datei aus", type=["xlsx"])

standard_zielwerte = [1200.0, 1100.0, 1300.0, 1000.0]
for idx, standardwert in enumerate(standard_zielwerte, start=1):
    st.session_state.setdefault(f"ziel_j{idx}", standardwert)

datei_bytes = uploaded_file.getvalue() if uploaded_file is not None else None
datei_hash = hashlib.sha256(datei_bytes).hexdigest() if datei_bytes is not None else None

# Zielwerte nur beim erstmaligen Laden einer neuen Datei übernehmen. Danach kann
# der Nutzer die Eingabefelder weiterhin manuell ändern.
if datei_hash is not None and st.session_state.get("geladene_datei_hash") != datei_hash:
    try:
        excel_datei = pd.ExcelFile(io.BytesIO(datei_bytes))
        if "Zielwerte" in excel_datei.sheet_names:
            zielwerte_df = pd.read_excel(excel_datei, sheet_name="Zielwerte")
            if "Zielwert" in zielwerte_df.columns:
                uebernommen = 0
                for idx in range(min(4, len(zielwerte_df))):
                    wert = zielwerte_df.iloc[idx]["Zielwert"]
                    if pd.notna(wert):
                        try:
                            st.session_state[f"ziel_j{idx + 1}"] = float(wert)
                            uebernommen += 1
                        except (TypeError, ValueError):
                            st.warning(f"Der Excel-Zielwert für Jahr {idx + 1} ist ungültig und wurde ignoriert.")
                if uebernommen:
                    st.success(f"✅ {uebernommen} Zielwerte aus dem Excel-Blatt ‚Zielwerte‘ übernommen.")
        st.session_state["geladene_datei_hash"] = datei_hash
    except Exception as e:
        st.warning(f"Die Zielwerte konnten nicht ausgelesen werden: {e}")
elif datei_hash is None:
    st.session_state.pop("geladene_datei_hash", None)

st.sidebar.header("2. Gewünschte Zielsummen")
ziel_j1 = st.sidebar.number_input("Ziel Gesamtsumme Jahr 1", step=1.0, key="ziel_j1")
ziel_j2 = st.sidebar.number_input("Ziel Gesamtsumme Jahr 2", step=1.0, key="ziel_j2")
ziel_j3 = st.sidebar.number_input("Ziel Gesamtsumme Jahr 3", step=1.0, key="ziel_j3")
ziel_j4 = st.sidebar.number_input("Ziel Gesamtsumme Jahr 4", step=1.0, key="ziel_j4")

st.sidebar.header("3. Berechnung")
anzahl_varianten = st.sidebar.number_input(
    "Anzahl der Varianten", min_value=1, max_value=50, value=10, step=1
)
zeit_pro_variante = st.sidebar.number_input(
    "Max. Zeit je Variante (Sekunden)", min_value=1, max_value=300, value=15, step=1
)

if uploaded_file is not None:
    try:
        excel_datei = pd.ExcelFile(io.BytesIO(datei_bytes))
        daten_blatt = "Daten" if "Daten" in excel_datei.sheet_names else excel_datei.sheet_names[0]
        df = pd.read_excel(excel_datei, sheet_name=daten_blatt)
        erforderliche_spalten = ["Objektname", "Jahr_1", "Jahr_2", "Jahr_3", "Jahr_4"]
        
        if not all(spalte in df.columns for spalte in erforderliche_spalten):
            st.error(f"❌ Fehler: Die Excel-Datei muss exakt die Spalten enthalten: {', '.join(erforderliche_spalten)}")
        else:
            df = df[erforderliche_spalten].copy()
            if df.empty:
                st.error("❌ Fehler: Die Excel-Datei enthält keine Datenzeilen.")
                st.stop()
            df["Objektname"] = df["Objektname"].astype(str).str.strip()
            doppelte_namen = df["Objektname"].duplicated(keep=False)
            ungueltige_zahlen = df[["Jahr_1", "Jahr_2", "Jahr_3", "Jahr_4"]].apply(
                lambda spalte: pd.to_numeric(spalte, errors="coerce")
            )

            if df["Objektname"].eq("").any() or df["Objektname"].eq("nan").any():
                st.error("❌ Fehler: Jeder Datensatz benötigt einen Objektname.")
                st.stop()
            if doppelte_namen.any():
                namen = ", ".join(df.loc[doppelte_namen, "Objektname"].unique())
                st.error(f"❌ Fehler: Objektnamen müssen eindeutig sein. Doppelt vorhanden: {namen}")
                st.stop()
            if ungueltige_zahlen.isna().any().any():
                st.error("❌ Fehler: Die vier Jahresspalten dürfen nur Zahlen enthalten.")
                st.stop()

            df[["Jahr_1", "Jahr_2", "Jahr_3", "Jahr_4"]] = ungueltige_zahlen
            st.success(f"✅ Excel-Struktur erfolgreich erkannt! ({len(df)} Zeilen gefunden)")
            
            if st.button("🚀 Top 10 Kombinationen berechnen"):
                objekte = df["Objektname"].tolist()
                j1_werte = dict(zip(df["Objektname"], df["Jahr_1"]))
                j2_werte = dict(zip(df["Objektname"], df["Jahr_2"]))
                j3_werte = dict(zip(df["Objektname"], df["Jahr_3"]))
                j4_werte = dict(zip(df["Objektname"], df["Jahr_4"]))
                
                gefundene_varianten = []
                varianten_metriken = []
                varianten_details = []
                anzahl_varianten_int = int(anzahl_varianten)
                zeit_pro_variante_int = int(zeit_pro_variante)
                berechnung_start = time.perf_counter()
                
                # Fortschrittsanzeige für den Nutzer
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for v_idx in range(1, anzahl_varianten_int + 1):
                    status_text.text(
                        f"Berechne Variante {v_idx} von {anzahl_varianten_int}... "
                        f"(Max. {zeit_pro_variante_int} Sek. pro Variante)"
                    )
                    progress_bar.progress((v_idx - 1) / anzahl_varianten_int)
                    
                    prob = pulp.LpProblem(f"Abweichung_Variante_{v_idx}", pulp.LpMinimize)
                    waehle = pulp.LpVariable.dicts("Waehle", objekte, cat='Binary')
                    
                    abw_pos_j1 = pulp.LpVariable("Abw_Pos_J1", lowBound=0)
                    abw_neg_j1 = pulp.LpVariable("Abw_Neg_J1", lowBound=0)
                    abw_pos_j2 = pulp.LpVariable("Abw_Pos_J2", lowBound=0)
                    abw_neg_j2 = pulp.LpVariable("Abw_Neg_J2", lowBound=0)
                    abw_pos_j3 = pulp.LpVariable("Abw_Pos_J3", lowBound=0)
                    abw_neg_j3 = pulp.LpVariable("Abw_Neg_J3", lowBound=0)
                    abw_pos_j4 = pulp.LpVariable("Abw_Pos_J4", lowBound=0)
                    abw_neg_j4 = pulp.LpVariable("Abw_Neg_J4", lowBound=0)
                    
                    prob += (abw_pos_j1 + abw_neg_j1 + abw_pos_j2 + abw_neg_j2 + 
                             abw_pos_j3 + abw_neg_j3 + abw_pos_j4 + abw_neg_j4)
                    
                    prob += pulp.lpSum([j1_werte[obj] * waehle[obj] for obj in objekte]) - abw_pos_j1 + abw_neg_j1 == ziel_j1
                    prob += pulp.lpSum([j2_werte[obj] * waehle[obj] for obj in objekte]) - abw_pos_j2 + abw_neg_j2 == ziel_j2
                    prob += pulp.lpSum([j3_werte[obj] * waehle[obj] for obj in objekte]) - abw_pos_j3 + abw_neg_j3 == ziel_j3
                    prob += pulp.lpSum([j4_werte[obj] * waehle[obj] for obj in objekte]) - abw_pos_j4 + abw_neg_j4 == ziel_j4
                    
                    for vorherige_sol in gefundene_varianten:
                        prob += pulp.lpSum([waehle[obj] if vorherige_sol[obj] == 1 else (1 - waehle[obj]) for obj in objekte]) <= len(objekte) - 1
                    
                    prob.solve(pulp.PULP_CBC_CMD(timeLimit=zeit_pro_variante_int, msg=False))
                    
                    solver_status = pulp.LpStatus[prob.status]
                    if solver_status in ["Optimal", "Feasible"] and waehle[objekte[0]].varValue is not None:
                        aktuelle_loesung = {obj: int(round(waehle[obj].varValue)) for obj in objekte}
                        gefundene_varianten.append(aktuelle_loesung)
                        
                        sum_j1 = sum(j1_werte[obj] * aktuelle_loesung[obj] for obj in objekte)
                        sum_j2 = sum(j2_werte[obj] * aktuelle_loesung[obj] for obj in objekte)
                        sum_j3 = sum(j3_werte[obj] * aktuelle_loesung[obj] for obj in objekte)
                        sum_j4 = sum(j4_werte[obj] * aktuelle_loesung[obj] for obj in objekte)
                        
                        total_abw = abs(sum_j1 - ziel_j1) + abs(sum_j2 - ziel_j2) + abs(sum_j3 - ziel_j3) + abs(sum_j4 - ziel_j4)
                        anzahl_ausgewaehlt = sum(aktuelle_loesung.values())
                        
                        varianten_metriken.append({
                            "Variante": f"Variante_{v_idx}",
                            "Status": "Bewiesenes Optimum" if solver_status == "Optimal" else "Gute Näherung (Zeitlimit)",
                            "Ausgewählte Objekte": anzahl_ausgewaehlt,
                            "Abweichung Gesamt": total_abw,
                            "Abweichung Jahr 1": sum_j1 - ziel_j1,
                            "Abweichung Jahr 2": sum_j2 - ziel_j2,
                            "Abweichung Jahr 3": sum_j3 - ziel_j3,
                            "Abweichung Jahr 4": sum_j4 - ziel_j4
                        })
                        varianten_details.append({
                            "Variante": f"Variante_{v_idx}",
                            "Summen": [sum_j1, sum_j2, sum_j3, sum_j4],
                            "Zielwerte": [ziel_j1, ziel_j2, ziel_j3, ziel_j4]
                        })
                    else:
                        break
                
                progress_bar.progress(1.0)
                status_text.text("Berechnung abgeschlossen!")
                berechnungsdauer = time.perf_counter() - berechnung_start
                
                # ==========================================
                # 4. AUSGABE DER ERGEBNISSE
                # ==========================================
                if len(gefundene_varianten) == 0:
                    st.error("Es konnte leider keine sinnvolle Kombination berechnet werden. Bitte prüfen Sie Ihre Daten.")
                else:
                    st.success(f"Erfolgreich {len(gefundene_varianten)} Varianten berechnet!")
                    
                    st.write("### 📈 Übersicht der berechneten Varianten:")
                    st.caption(
                        "Abweichung je Jahr = berechnete Summe − Zielwert (+ über Ziel, − unter Ziel). "
                        "🟢 bis 1 %, 🟡 bis 5 %, 🔴 über 5 % des jeweiligen Zielwerts."
                    )
                    metriken_df = pd.DataFrame(varianten_metriken)
                    metriken_format = {
                        spalte: st.column_config.NumberColumn(format="%.1f")
                        for spalte in [
                            "Abweichung Gesamt", "Abweichung Jahr 1", "Abweichung Jahr 2",
                            "Abweichung Jahr 3", "Abweichung Jahr 4"
                        ]
                    }
                    zielwerte_fuer_spalten = {
                        f"Abweichung Jahr {idx + 1}": abs(zielwert)
                        for idx, zielwert in enumerate([ziel_j1, ziel_j2, ziel_j3, ziel_j4])
                    }

                    def faerbe_abweichungen(spalte):
                        zielwert = zielwerte_fuer_spalten.get(spalte.name)
                        if zielwert is None:
                            return [""] * len(spalte)
                        basis = zielwert if zielwert > 0 else 1
                        farben = []
                        for wert in spalte:
                            anteil = abs(wert) / basis
                            if anteil <= 0.01:
                                farben.append("background-color: #d9ead3; color: #1b4332")
                            elif anteil <= 0.05:
                                farben.append("background-color: #fff2cc; color: #6b4f00")
                            else:
                                farben.append("background-color: #f4cccc; color: #7a1f1f")
                        return farben

                    metriken_stil = metriken_df.style.apply(
                        faerbe_abweichungen,
                        subset=list(zielwerte_fuer_spalten),
                        axis=0
                    )
                    st.dataframe(
                        metriken_stil,
                        use_container_width=True,
                        column_config=metriken_format,
                        hide_index=True
                    )
                    
                    # Spalten für alle gefundenen Varianten anhängen
                    for idx, sol in enumerate(gefundene_varianten):
                        v_name = f"Variante_{idx+1} (1=Ja)"
                        df[v_name] = [sol[obj] for obj in df["Objektname"]]
                    
                    st.write("### 📋 Daten-Vorschau inklusive Varianten-Spalten:")
                    jahres_format = {
                        spalte: st.column_config.NumberColumn(format="%.1f")
                        for spalte in ["Jahr_1", "Jahr_2", "Jahr_3", "Jahr_4"]
                    }
                    st.dataframe(df, column_config=jahres_format, hide_index=True)
                    
                    # ==========================================
                    # 5. EXCEL EXPORT
                    # ==========================================
                    vergleich_zeilen = []
                    for detail in varianten_details:
                        for jahr_idx, (zielwert, summe) in enumerate(
                            zip(detail["Zielwerte"], detail["Summen"]), start=1
                        ):
                            vergleich_zeilen.append({
                                "Variante": detail["Variante"],
                                "Jahr": f"Jahr {jahr_idx}",
                                "Zielwert": zielwert,
                                "Berechnete Summe": summe,
                                "Abweichung": summe - zielwert
                            })
                    ziel_ist_df = pd.DataFrame(vergleich_zeilen)
                    parameter_df = pd.DataFrame({
                        "Parameter": [
                            "Zielwert Jahr 1", "Zielwert Jahr 2", "Zielwert Jahr 3", "Zielwert Jahr 4",
                            "Angeforderte Varianten", "Berechnete Varianten",
                            "Zeitlimit je Variante (Sek.)", "Gesamte Rechenzeit (Sek.)"
                        ],
                        "Wert": [
                            ziel_j1, ziel_j2, ziel_j3, ziel_j4,
                            anzahl_varianten_int, len(gefundene_varianten),
                            zeit_pro_variante_int, round(berechnungsdauer, 1)
                        ]
                    })

                    buffer_export = io.BytesIO()
                    with pd.ExcelWriter(buffer_export, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name="Objekt_Auswahl")
                        metriken_df.to_excel(writer, index=False, sheet_name="Varianten_Vergleich")
                        ziel_ist_df.to_excel(writer, index=False, sheet_name="Ziel_Ist_Vergleich")
                        parameter_df.to_excel(writer, index=False, sheet_name="Parameter")

                        for idx, sol in enumerate(gefundene_varianten, start=1):
                            varianten_df = df[erforderliche_spalten].copy()
                            varianten_df["Ausgewählt"] = [sol[obj] for obj in df["Objektname"]]
                            varianten_df.to_excel(writer, index=False, sheet_name=f"Variante_{idx}")

                        kopf_fuellung = PatternFill("solid", fgColor="1F4E78")
                        kopf_schrift = Font(color="FFFFFF", bold=True)
                        auswahl_fuellung = PatternFill("solid", fgColor="C6EFCE")

                        for tabellenblatt in writer.book.worksheets:
                            tabellenblatt.freeze_panes = "A2"
                            tabellenblatt.auto_filter.ref = tabellenblatt.dimensions
                            for zelle in tabellenblatt[1]:
                                zelle.fill = kopf_fuellung
                                zelle.font = kopf_schrift
                                zelle.alignment = Alignment(horizontal="center")
                            for spalten_idx, spalte in enumerate(tabellenblatt.columns, start=1):
                                breite = min(max(len(str(zelle.value or "")) for zelle in spalte) + 2, 35)
                                tabellenblatt.column_dimensions[get_column_letter(spalten_idx)].width = breite
                            for zeile in tabellenblatt.iter_rows(min_row=2):
                                for zelle in zeile:
                                    if isinstance(zelle.value, float):
                                        zelle.number_format = "0.0"

                        objekt_blatt = writer.book["Objekt_Auswahl"]
                        for spalten_idx in range(6, objekt_blatt.max_column + 1):
                            for zeilen_idx in range(2, objekt_blatt.max_row + 1):
                                zelle = objekt_blatt.cell(zeilen_idx, spalten_idx)
                                if zelle.value == 1:
                                    zelle.fill = auswahl_fuellung

                        for idx in range(1, len(gefundene_varianten) + 1):
                            varianten_blatt = writer.book[f"Variante_{idx}"]
                            for zeilen_idx in range(2, varianten_blatt.max_row + 1):
                                if varianten_blatt.cell(zeilen_idx, 6).value == 1:
                                    for spalten_idx in range(1, varianten_blatt.max_column + 1):
                                        varianten_blatt.cell(zeilen_idx, spalten_idx).fill = auswahl_fuellung
                        
                    st.download_button(
                        label="📥 Excel mit berechneten Varianten herunterladen",
                        data=buffer_export.getvalue(),
                        file_name="4d_optimierung_varianten.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
    except Exception as e:
        st.error(f"Fehler bei der Verarbeitung: {e}")
