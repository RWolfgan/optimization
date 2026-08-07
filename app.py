import io
import time

import pandas as pd
import pulp
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="SuSa-zu-BWA Optimierer", page_icon="📊", layout="wide")
st.title("📊 SuSa-zu-BWA Optimierer")
st.write(
    "Ordnet passende Konten höchstens einer BWA-Position zu und minimiert dabei die "
    "Abweichungen von den Zielwerten über alle gemeinsamen Jahre. Nicht passende Konten "
    "dürfen unzugeordnet bleiben."
)

# Diese Spalten sind fest vorgegeben. Jede weitere Spalte wird als Jahr bzw.
# frei benannte Vergleichsperiode interpretiert.
KONTEN_STAMMSPALTEN = ["Konto", "Kontobezeichnung"]
BWA_STAMMSPALTEN = ["BWA_Position", "BWA_Bezeichnung"]


def vorlagen_download(konten_df, bwa_df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        konten_df.to_excel(writer, index=False, sheet_name="Konten")
        bwa_df.to_excel(writer, index=False, sheet_name="BWA")
        formatiere_arbeitsmappe(writer)
    return buffer.getvalue()


def lese_excel(upload, bevorzugtes_blatt):
    """Liest ein fest benanntes Tabellenblatt aus der hochgeladenen Arbeitsmappe."""
    datei = pd.ExcelFile(io.BytesIO(upload.getvalue()))
    if bevorzugtes_blatt not in datei.sheet_names:
        raise ValueError(f"Das Tabellenblatt „{bevorzugtes_blatt}“ fehlt.")
    df = pd.read_excel(datei, sheet_name=bevorzugtes_blatt).dropna(how="all")
    df.columns = [str(spalte).strip() for spalte in df.columns]
    if df.columns.duplicated().any():
        raise ValueError("Die Excel-Datei enthält doppelte Spaltennamen.")
    return df


def kennung_als_text(wert):
    """Erhält Konten wie 1000 als '1000' statt als '1000.0'."""
    if pd.isna(wert):
        return ""
    if isinstance(wert, float) and wert.is_integer():
        return str(int(wert))
    return str(wert).strip()


def validiere_tabelle(df, stammspalten, tabellenname):
    """Prüft Pflichtspalten, eindeutige Kennungen und numerische Periodenwerte."""
    fehlend = [spalte for spalte in stammspalten if spalte not in df.columns]
    if fehlend:
        raise ValueError(f"{tabellenname}: Fehlende Spalten: {', '.join(fehlend)}")
    if df.empty:
        raise ValueError(f"{tabellenname}: Die Datei enthält keine Datenzeilen.")

    df = df.copy()
    df[stammspalten[0]] = df[stammspalten[0]].map(kennung_als_text)
    df[stammspalten[1]] = df[stammspalten[1]].fillna("").astype(str).str.strip()
    if df[stammspalten[0]].eq("").any():
        raise ValueError(f"{tabellenname}: Jede Zeile benötigt einen Wert in „{stammspalten[0]}“.")
    if df[stammspalten[0]].duplicated().any():
        doppelt = ", ".join(df.loc[df[stammspalten[0]].duplicated(False), stammspalten[0]].unique())
        raise ValueError(f"{tabellenname}: Diese Kennungen sind doppelt vorhanden: {doppelt}")

    jahre = [spalte for spalte in df.columns if spalte not in stammspalten]
    if not jahre:
        raise ValueError(f"{tabellenname}: Es wurde keine Jahresspalte gefunden.")
    for jahr in jahre:
        df[jahr] = pd.to_numeric(df[jahr], errors="coerce")
        if df[jahr].isna().any():
            fehlerzeilen = ", ".join(str(i + 2) for i in df.index[df[jahr].isna()][:10])
            raise ValueError(f"{tabellenname}: Ungültiger Wert in „{jahr}“, Zeile(n) {fehlerzeilen}.")
    return df, jahre


def formatiere_arbeitsmappe(writer):
    """Wendet ein einheitliches, gut lesbares Format auf alle Exportblätter an."""
    kopf_fuellung = PatternFill("solid", fgColor="1F4E78")
    kopf_schrift = Font(color="FFFFFF", bold=True)
    for blatt in writer.book.worksheets:
        blatt.freeze_panes = "A2"
        blatt.auto_filter.ref = blatt.dimensions
        for zelle in blatt[1]:
            zelle.fill = kopf_fuellung
            zelle.font = kopf_schrift
            zelle.alignment = Alignment(horizontal="center")
        for index, spalte in enumerate(blatt.columns, start=1):
            breite = min(max(len(str(zelle.value or "")) for zelle in spalte) + 2, 35)
            blatt.column_dimensions[get_column_letter(index)].width = breite
        for zeile in blatt.iter_rows(min_row=2):
            for zelle in zeile:
                if isinstance(zelle.value, float):
                    zelle.number_format = "0.0"


st.sidebar.header("1. Excel-Vorlagen")
beispiel_jahre = ["2024", "2025", "2026"]
konten_vorlage = pd.DataFrame({
    "Konto": [1000, 1200, 4000, 5000],
    "Kontobezeichnung": ["Kasse", "Bank", "Umsatzerlöse", "Materialaufwand"],
    **{jahr: [0.0] * 4 for jahr in beispiel_jahre},
})
bwa_vorlage = pd.DataFrame({
    "BWA_Position": ["P01", "P02", "P03"],
    "BWA_Bezeichnung": ["Umsatzerlöse", "Materialaufwand", "Sonstige Position"],
    **{jahr: [0.0] * 3 for jahr in beispiel_jahre},
})
st.sidebar.download_button(
    "📄 Excel-Vorlage herunterladen",
    vorlagen_download(konten_vorlage, bwa_vorlage),
    "susa_bwa_vorlage.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.sidebar.header("2. Berechnung")
zeitlimit = st.sidebar.number_input(
    "Maximale Rechenzeit (Sekunden)", min_value=5, max_value=1800, value=120, step=5
)

excel_upload = st.file_uploader(
    "Excel-Datei mit den Tabellenblättern „Konten“ und „BWA“",
    type=["xlsx"],
    key="susa_bwa_upload",
)

if excel_upload is None:
    st.info(
        "Bitte die ausgefüllte Excel-Datei auswählen. Sie muss die Tabellenblätter „Konten“ "
        "und „BWA“ mit identisch benannten Jahresspalten enthalten."
    )
    st.stop()

try:
    konten_roh = lese_excel(excel_upload, "Konten")
    bwa_roh = lese_excel(excel_upload, "BWA")
    konten_df, konten_jahre = validiere_tabelle(konten_roh, KONTEN_STAMMSPALTEN, "Konten-Datei")
    bwa_df, bwa_jahre = validiere_tabelle(bwa_roh, BWA_STAMMSPALTEN, "BWA-Datei")
    # Nur Perioden mit identischen Namen lassen sich sicher miteinander vergleichen.
    if set(konten_jahre) != set(bwa_jahre):
        nur_konten = sorted(set(konten_jahre) - set(bwa_jahre))
        nur_bwa = sorted(set(bwa_jahre) - set(konten_jahre))
        details = []
        if nur_konten:
            details.append(f"nur in Konten: {', '.join(nur_konten)}")
        if nur_bwa:
            details.append(f"nur in BWA: {', '.join(nur_bwa)}")
        raise ValueError("Die Jahresspalten stimmen nicht überein (" + "; ".join(details) + ").")
    jahre = konten_jahre
except Exception as error:
    st.error(f"❌ {error}")
    st.stop()

anzahl_variablen = len(konten_df) * len(bwa_df)
st.success(
    f"✅ Eingaben erkannt: {len(konten_df)} Konten, {len(bwa_df)} BWA-Positionen, "
    f"{len(jahre)} Jahre ({anzahl_variablen:,} Zuordnungsvariablen)."
)

with st.expander("Eingabedaten prüfen"):
    st.write("**Konten**")
    st.dataframe(konten_df, hide_index=True, use_container_width=True)
    st.write("**BWA-Ziele**")
    st.dataframe(bwa_df, hide_index=True, use_container_width=True)

if st.button("🚀 Konten optimal zuordnen", type="primary"):
    startzeit = time.perf_counter()
    with st.spinner("CBC optimiert die Zuordnung …"):
        # Ein globales MILP-Modell verhindert, dass eine früh bearbeitete BWA-Zeile
        # gute Konten verbraucht, die an einer späteren Position dringender benötigt werden.
        problem = pulp.LpProblem("SuSa_zu_BWA", pulp.LpMinimize)
        konten_indizes = range(len(konten_df))
        bwa_indizes = range(len(bwa_df))
        # zuordnung[konto][position] ist 1, wenn das Konto dieser BWA-Position
        # zugeordnet wird. Es gibt bewusst keine separate Variable pro Jahr:
        # Die einmal gewählte Position gilt dadurch automatisch für alle Jahre.
        zuordnung = pulp.LpVariable.dicts(
            "Zuordnung", (konten_indizes, bwa_indizes), cat="Binary"
        )
        # Positive und negative Abweichungsvariablen bilden den Absolutbetrag
        # linear ab, damit CBC die Zielabweichungen minimieren kann.
        abw_pos = pulp.LpVariable.dicts(
            "Abweichung_Pos", (bwa_indizes, range(len(jahre))), lowBound=0
        )
        abw_neg = pulp.LpVariable.dicts(
            "Abweichung_Neg", (bwa_indizes, range(len(jahre))), lowBound=0
        )

        # Zielfunktion: Summe aller absoluten Abweichungen über sämtliche
        # BWA-Positionen und Jahre minimieren.
        problem += pulp.lpSum(
            abw_pos[p][j] + abw_neg[p][j]
            for p in bwa_indizes for j in range(len(jahre))
        )

        # Ein Konto darf höchstens einer Position angehören. Durch "<= 1" darf
        # es unzugeordnet bleiben; "== 1" würde eine Zuordnung erzwingen.
        for konto_idx in konten_indizes:
            problem += pulp.lpSum(zuordnung[konto_idx][p] for p in bwa_indizes) <= 1

        # Für jede Position und jedes Jahr wird die Summe der zugeordneten
        # Konten mit dem eingegebenen BWA-Zielwert verglichen.
        for position_idx in bwa_indizes:
            for jahr_idx, jahr in enumerate(jahre):
                ist_summe = pulp.lpSum(
                    float(konten_df.iloc[konto_idx][jahr]) * zuordnung[konto_idx][position_idx]
                    for konto_idx in konten_indizes
                )
                zielwert = float(bwa_df.iloc[position_idx][jahr])
                problem += (
                    ist_summe - abw_pos[position_idx][jahr_idx]
                    + abw_neg[position_idx][jahr_idx] == zielwert
                )

        problem.solve(pulp.PULP_CBC_CMD(timeLimit=int(zeitlimit), msg=False))

    status = pulp.LpStatus[problem.status]
    if status not in ["Optimal", "Feasible"] or zuordnung[0][0].varValue is None:
        st.error(f"Keine verwertbare Zuordnung gefunden. Solver-Status: {status}")
        st.stop()

    laufzeit = time.perf_counter() - startzeit
    # Binäre Solverwerte werden wieder in die fachlichen BWA-Kennungen übersetzt.
    zugeordnete_positionen = []
    for konto_idx in konten_indizes:
        ausgewaehlt = [
            p for p in bwa_indizes
            if (zuordnung[konto_idx][p].varValue or 0) > 0.5
        ]
        zugeordnete_positionen.append(ausgewaehlt[0] if ausgewaehlt else None)

    zuordnung_df = konten_df.copy()
    zuordnung_df["BWA_Position"] = [
        bwa_df.iloc[p]["BWA_Position"] if p is not None else ""
        for p in zugeordnete_positionen
    ]
    zuordnung_df["BWA_Bezeichnung"] = [
        bwa_df.iloc[p]["BWA_Bezeichnung"] if p is not None else "Nicht zugeordnet"
        for p in zugeordnete_positionen
    ]
    anzahl_nicht_zugeordnet = sum(p is None for p in zugeordnete_positionen)

    # Ergebnisaggregation je BWA-Position: Ziel, Ist und signierte Abweichung.
    vergleich_zeilen = []
    for position_idx in bwa_indizes:
        konto_mask = [p == position_idx for p in zugeordnete_positionen]
        zeile = {
            "BWA_Position": bwa_df.iloc[position_idx]["BWA_Position"],
            "BWA_Bezeichnung": bwa_df.iloc[position_idx]["BWA_Bezeichnung"],
            "Anzahl Konten": sum(konto_mask),
        }
        gesamt_abweichung = 0.0
        for jahr in jahre:
            zielwert = float(bwa_df.iloc[position_idx][jahr])
            istwert = float(konten_df.loc[konto_mask, jahr].sum())
            abweichung = istwert - zielwert
            zeile[f"{jahr} Ziel"] = zielwert
            zeile[f"{jahr} Ist"] = istwert
            zeile[f"{jahr} Abweichung"] = abweichung
            gesamt_abweichung += abs(abweichung)
        zeile["Abweichung Gesamt"] = gesamt_abweichung
        vergleich_zeilen.append(zeile)

    vergleich_df = pd.DataFrame(vergleich_zeilen)
    gesamt_abweichung = vergleich_df["Abweichung Gesamt"].sum()
    st.success(
        f"Zuordnung abgeschlossen: {status}, Gesamtabweichung {gesamt_abweichung:,.1f}, "
        f"{anzahl_nicht_zugeordnet} Konten nicht zugeordnet, Rechenzeit {laufzeit:.1f} Sekunden."
    )

    st.write("### BWA-Zielvergleich")
    st.dataframe(
        vergleich_df,
        hide_index=True,
        use_container_width=True,
        column_config={
            spalte: st.column_config.NumberColumn(format="%.1f")
            for spalte in vergleich_df.columns
            if spalte.endswith((" Ziel", " Ist", " Abweichung")) or spalte == "Abweichung Gesamt"
        },
    )
    st.write("### Kontenzuordnung")
    st.dataframe(
        zuordnung_df,
        hide_index=True,
        use_container_width=True,
        column_config={jahr: st.column_config.NumberColumn(format="%.1f") for jahr in jahre},
    )

    if anzahl_nicht_zugeordnet:
        st.write("### Nicht zugeordnete Konten")
        st.dataframe(
            zuordnung_df[zuordnung_df["BWA_Bezeichnung"] == "Nicht zugeordnet"],
            hide_index=True,
            use_container_width=True,
            column_config={jahr: st.column_config.NumberColumn(format="%.1f") for jahr in jahre},
        )

    parameter_df = pd.DataFrame({
        "Parameter": [
            "Solver-Status", "Zeitlimit (Sek.)", "Rechenzeit (Sek.)",
            "Gesamtabweichung", "Konten gesamt", "Konten nicht zugeordnet"
        ],
        "Wert": [
            status, int(zeitlimit), round(laufzeit, 1), round(gesamt_abweichung, 1),
            len(konten_df), anzahl_nicht_zugeordnet
        ],
    })
    nicht_zugeordnet_df = zuordnung_df[zuordnung_df["BWA_Bezeichnung"] == "Nicht zugeordnet"].copy()
    export_buffer = io.BytesIO()
    with pd.ExcelWriter(export_buffer, engine="openpyxl") as writer:
        zuordnung_df.to_excel(writer, index=False, sheet_name="Kontenzuordnung")
        vergleich_df.to_excel(writer, index=False, sheet_name="BWA_Vergleich")
        bwa_df.to_excel(writer, index=False, sheet_name="BWA_Ziele")
        parameter_df.to_excel(writer, index=False, sheet_name="Parameter")
        nicht_zugeordnet_df.to_excel(writer, index=False, sheet_name="Nicht_zugeordnet")
        formatiere_arbeitsmappe(writer)

    st.download_button(
        "📥 Zuordnung als Excel herunterladen",
        export_buffer.getvalue(),
        "susa_bwa_zuordnung.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
